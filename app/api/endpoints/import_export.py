"""
Import/Export endpoints for OTOBO integration
"""

import io
import csv
import json
import xml.etree.ElementTree as ET
import socket
import platform
from datetime import datetime
from typing import List
from uuid import UUID, uuid4
import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import Response, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from app.core.database import get_db
from app.core.config import settings
from app.models.asset import Asset, AssetType, AssetTypeEnum
from app.schemas.common import ExportFormat, ImportResult

router = APIRouter()


@router.post("/import/csv", response_model=ImportResult)
async def import_csv(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Import assets from CSV file"""
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")
    
    contents = await file.read()
    content_str = contents.decode('utf-8')
    
    # Skip comment lines (metadata) if present
    lines = content_str.split('\n')
    clean_lines = []
    for line in lines:
        if not line.startswith('#'):
            clean_lines.append(line)
    
    # Rejoin the clean lines
    clean_content = '\n'.join(clean_lines)
    
    # Read the CSV
    try:
        df = pd.read_csv(io.StringIO(clean_content))
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"Error parsing CSV file: {str(e)}"
        )
    
    # Check if dataframe is empty
    if df.empty:
        raise HTTPException(
            status_code=400,
            detail="CSV file is empty or contains no valid data"
        )
    
    # Required columns
    required_columns = ['name', 'asset_type', 'parent_name']
    missing_columns = [col for col in required_columns if col not in df.columns]
    if missing_columns:
        raise HTTPException(
            status_code=400,
            detail=f"CSV must contain columns: {missing_columns}. Found columns: {list(df.columns)}"
        )
    
    # Get asset types mapping
    result = await db.execute(select(AssetType))
    asset_types = {at.name: at for at in result.scalars().all()}
    
    imported = 0
    failed = 0
    errors = []
    warnings = []
    
    # Sort dataframe so that assets with no parent come first
    # This helps ensure parents are created before children
    df['has_parent'] = df['parent_name'].notna() & (df['parent_name'] != '')
    df = df.sort_values('has_parent')
    df = df.reset_index(drop=True)
    
    # Process each row
    for idx, row in df.iterrows():
        try:
            # Skip if asset name is empty
            if pd.isna(row.get('name')) or not str(row['name']).strip():
                errors.append({
                    "row": idx + 1,
                    "error": "Asset name is required"
                })
                failed += 1
                continue
            
            # Find asset type
            asset_type = asset_types.get(row['asset_type'])
            if not asset_type:
                errors.append({
                    "row": idx + 1,
                    "error": f"Unknown asset type: {row['asset_type']}"
                })
                failed += 1
                continue
            
            # Find parent if specified
            parent_id = None
            if pd.notna(row.get('parent_name')) and row['parent_name']:
                parent_result = await db.execute(
                    select(Asset).where(Asset.name == row['parent_name'])
                )
                parent = parent_result.scalar_one_or_none()
                if parent:
                    parent_id = parent.id
                else:
                    # Parent not found - log warning but continue
                    warnings.append({
                        "row": idx + 1,
                        "message": f"Parent '{row['parent_name']}' not found, asset created at root level"
                    })
            
            # Check if asset already exists (by name and parent)
            existing_check = await db.execute(
                select(Asset).where(
                    Asset.name == row['name'],
                    Asset.parent_id == parent_id
                )
            )
            existing_asset = existing_check.scalar_one_or_none()
            
            if existing_asset:
                # Update existing asset instead of creating duplicate
                existing_asset.description = row.get('description') if pd.notna(row.get('description')) else existing_asset.description
                existing_asset.asset_type_id = asset_type.id
                existing_asset.status = row.get('status', 'active') if pd.notna(row.get('status')) else existing_asset.status
                existing_asset.external_id = row.get('external_id') if pd.notna(row.get('external_id')) else existing_asset.external_id
                existing_asset.external_system = row.get('external_system', 'OTOBO') if pd.notna(row.get('external_system')) else existing_asset.external_system
                existing_asset.version = row.get('version') if pd.notna(row.get('version')) else existing_asset.version
                
                # Handle properties - parse JSON if provided
                if pd.notna(row.get('properties')) and row.get('properties'):
                    try:
                        existing_asset.properties = json.loads(row['properties'])
                    except:
                        existing_asset.properties = {}
                
                # Handle tags
                if pd.notna(row.get('tags')) and row.get('tags'):
                    existing_asset.tags = [t.strip() for t in str(row['tags']).split(',') if t.strip()]
                
                imported += 1
            else:
                # Create new asset
                asset = Asset(
                    name=row['name'],
                    description=row.get('description') if pd.notna(row.get('description')) else '',
                    asset_type_id=asset_type.id,
                    parent_id=parent_id,
                    status=row.get('status', 'active') if pd.notna(row.get('status')) else 'active',
                    external_id=row.get('external_id') if pd.notna(row.get('external_id')) else None,
                    external_system=row.get('external_system', 'OTOBO') if pd.notna(row.get('external_system')) else 'OTOBO',
                    version=row.get('version') if pd.notna(row.get('version')) else None,
                    properties={},
                    tags=[]
                )
                
                # Handle properties
                if pd.notna(row.get('properties')) and row.get('properties'):
                    try:
                        asset.properties = json.loads(row['properties'])
                    except:
                        asset.properties = {}
                
                # Handle tags
                if pd.notna(row.get('tags')) and row.get('tags'):
                    asset.tags = [t.strip() for t in str(row['tags']).split(',') if t.strip()]
                
                # Set the asset_type relationship for URN generation
                asset.asset_type = asset_type
                base_urn = asset.generate_urn()
                
                # Check if URN already exists and make it unique if needed
                urn_check = await db.execute(
                    select(Asset).where(Asset.urn == base_urn)
                )
                existing_with_urn = urn_check.scalar_one_or_none()
                if existing_with_urn:
                    # URN exists - update the existing asset instead
                    existing_with_urn.description = asset.description
                    existing_with_urn.status = asset.status
                    existing_with_urn.version = asset.version
                    existing_with_urn.external_id = asset.external_id
                    existing_with_urn.properties = asset.properties
                    existing_with_urn.tags = asset.tags
                    imported += 1
                else:
                    asset.urn = base_urn
                    db.add(asset)
                    imported += 1
                    
                # Commit after each successful asset to avoid batch conflicts
                await db.flush()
            
        except Exception as e:
            errors.append({
                "row": idx + 1,
                "error": str(e)
            })
            failed += 1
    
    # Commit all changes
    await db.commit()
    
    return ImportResult(
        success=failed == 0,
        total_records=len(df),
        imported=imported,
        failed=failed,
        errors=errors,
        warnings=warnings
    )


@router.post("/import/json", response_model=ImportResult)
async def import_json(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db)
):
    """Import assets from JSON file"""
    if not file.filename.endswith('.json'):
        raise HTTPException(status_code=400, detail="File must be JSON")
    
    contents = await file.read()
    data = json.loads(contents)
    
    if not isinstance(data, list):
        data = [data]
    
    # Get asset types mapping
    result = await db.execute(select(AssetType))
    asset_types = {at.name: at for at in result.scalars().all()}
    
    imported = 0
    failed = 0
    errors = []
    
    # Create a mapping of names to IDs for parent resolution
    name_to_id = {}
    
    # First pass: create all assets without parents
    for idx, item in enumerate(data):
        try:
            asset_type = asset_types.get(item['asset_type'])
            if not asset_type:
                errors.append({
                    "index": idx,
                    "error": f"Unknown asset type: {item['asset_type']}"
                })
                failed += 1
                continue
            
            asset = Asset(
                name=item['name'],
                description=item.get('description'),
                asset_type_id=asset_type.id,
                parent_id=None,  # Set in second pass
                status=item.get('status', 'active'),
                external_id=item.get('external_id'),
                external_system=item.get('external_system', 'OTOBO'),
                version=item.get('version'),
                properties=item.get('properties', {}),
                tags=item.get('tags', [])
            )
            asset.urn = asset.generate_urn()
            
            db.add(asset)
            await db.flush()  # Get the ID
            name_to_id[asset.name] = asset.id
            imported += 1
            
        except Exception as e:
            errors.append({
                "index": idx,
                "error": str(e)
            })
            failed += 1
    
    # Second pass: set parent relationships
    for item in data:
        if item.get('parent_name') and item['name'] in name_to_id:
            parent_id = name_to_id.get(item['parent_name'])
            if parent_id:
                result = await db.execute(
                    select(Asset).where(Asset.name == item['name'])
                )
                asset = result.scalar_one_or_none()
                if asset:
                    asset.parent_id = parent_id
    
    await db.commit()
    
    return ImportResult(
        success=failed == 0,
        total_records=len(data),
        imported=imported,
        failed=failed,
        errors=errors
    )


@router.get("/export/{format}")
async def export_assets(
    format: ExportFormat,
    request: Request,
    db: AsyncSession = Depends(get_db)
):
    """Export all assets in specified format"""
    # Get export metadata
    export_metadata = {
        "export_date": datetime.utcnow().isoformat(),
        "app_version": settings.APP_VERSION,
        "hostname": socket.gethostname(),
        "host_ip": socket.gethostbyname(socket.gethostname()),
        "client_ip": request.client.host if request.client else "unknown",
        "platform": f"{platform.system()} {platform.release()}",
        "python_version": platform.python_version()
    }
    
    # Get all assets with their types and build hierarchy
    query = select(Asset).options(
        selectinload(Asset.asset_type), 
        selectinload(Asset.parent),
        selectinload(Asset.children)
    )
    result = await db.execute(query)
    all_assets = result.scalars().all()
    
    # Build hierarchical structure
    def build_hierarchy(parent_id=None, level=0):
        """Build hierarchical structure of assets"""
        assets_at_level = []
        for asset in all_assets:
            if asset.parent_id == parent_id:
                asset.hierarchy_level = level
                assets_at_level.append(asset)
                # Add children recursively
                children = build_hierarchy(asset.id, level + 1)
                assets_at_level.extend(children)
        return assets_at_level
    
    # Get hierarchically ordered assets
    ordered_assets = build_hierarchy()
    
    if format == ExportFormat.CSV:
        return export_to_csv(ordered_assets, export_metadata)
    elif format == ExportFormat.JSON:
        return export_to_json(ordered_assets, export_metadata)
    elif format == ExportFormat.XML:
        return export_to_xml_hierarchical(all_assets, export_metadata)
    elif format == ExportFormat.EXCEL:
        return export_to_excel_hierarchical(ordered_assets, export_metadata)
    else:
        raise HTTPException(status_code=400, detail="Unsupported format")


def export_to_csv(assets: List[Asset], metadata: dict = None) -> Response:
    """Export assets to CSV format with metadata header"""
    output = io.StringIO()
    
    # Write metadata as comments at the top
    if metadata:
        output.write(f"# AssetDNA Export\n")
        output.write(f"# Export Date: {metadata['export_date']}\n")
        output.write(f"# Application Version: {metadata['app_version']}\n")
        output.write(f"# Host: {metadata['hostname']} ({metadata['host_ip']})\n")
        output.write(f"# Client IP: {metadata['client_ip']}\n")
        output.write(f"# Platform: {metadata['platform']}\n")
        output.write(f"# Total Assets: {len(assets)}\n")
        output.write("#\n")
    
    writer = csv.DictWriter(
        output,
        fieldnames=[
            'id', 'urn', 'name', 'description', 'asset_type', 'parent_name',
            'status', 'lifecycle_stage', 'external_id', 'external_system',
            'version', 'properties', 'tags', 'created_at', 'updated_at'
        ]
    )
    writer.writeheader()
    
    for asset in assets:
        writer.writerow({
            'id': str(asset.id),
            'urn': asset.urn,
            'name': asset.name,
            'description': asset.description or '',
            'asset_type': asset.asset_type.name,
            'parent_name': asset.parent.name if asset.parent else '',
            'status': asset.status,
            'lifecycle_stage': asset.lifecycle_stage or '',
            'external_id': asset.external_id or '',
            'external_system': asset.external_system or '',
            'version': asset.version or '',
            'properties': json.dumps(asset.properties),
            'tags': ','.join(asset.tags) if asset.tags else '',
            'created_at': asset.created_at.isoformat(),
            'updated_at': asset.updated_at.isoformat()
        })
    
    return Response(
        content=output.getvalue(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=assets.csv"}
    )


def export_to_json(assets: List[Asset], metadata: dict = None) -> Response:
    """Export assets to JSON format with metadata"""
    export_data = {
        "metadata": metadata or {},
        "total_assets": len(assets),
        "assets": []
    }
    
    for asset in assets:
        export_data["assets"].append({
            'id': str(asset.id),
            'urn': asset.urn,
            'name': asset.name,
            'description': asset.description,
            'asset_type': asset.asset_type.name,
            'parent_name': asset.parent.name if asset.parent else None,
            'status': asset.status,
            'lifecycle_stage': asset.lifecycle_stage,
            'external_id': asset.external_id,
            'external_system': asset.external_system,
            'version': asset.version,
            'properties': asset.properties,
            'tags': asset.tags,
            'created_at': asset.created_at.isoformat(),
            'updated_at': asset.updated_at.isoformat()
        })
    
    return Response(
        content=json.dumps(export_data, indent=2),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=assets.json"}
    )


def export_to_xml(assets: List[Asset]) -> Response:
    """Export assets to XML format - flat structure"""
    import xml.dom.minidom as minidom
    
    root = ET.Element("assets")
    root.set("export_date", datetime.utcnow().isoformat())
    root.set("total_count", str(len(assets)))
    
    for asset in assets:
        asset_elem = ET.SubElement(root, "asset")
        asset_elem.set("id", str(asset.id))
        asset_elem.set("urn", asset.urn)
        
        ET.SubElement(asset_elem, "name").text = asset.name
        ET.SubElement(asset_elem, "description").text = asset.description or ""
        ET.SubElement(asset_elem, "asset_type").text = asset.asset_type.name
        ET.SubElement(asset_elem, "parent_name").text = asset.parent.name if asset.parent else ""
        ET.SubElement(asset_elem, "status").text = asset.status
        ET.SubElement(asset_elem, "lifecycle_stage").text = asset.lifecycle_stage or ""
        ET.SubElement(asset_elem, "external_id").text = asset.external_id or ""
        ET.SubElement(asset_elem, "external_system").text = asset.external_system or ""
        ET.SubElement(asset_elem, "version").text = asset.version or ""
        
        # Properties
        if asset.properties:
            props_elem = ET.SubElement(asset_elem, "properties")
            for key, value in asset.properties.items():
                prop_elem = ET.SubElement(props_elem, "property")
                prop_elem.set("key", key)
                prop_elem.text = str(value)
        
        # Tags
        if asset.tags:
            tags_elem = ET.SubElement(asset_elem, "tags")
            for tag in asset.tags:
                ET.SubElement(tags_elem, "tag").text = tag
        
        ET.SubElement(asset_elem, "created_at").text = asset.created_at.isoformat()
        ET.SubElement(asset_elem, "updated_at").text = asset.updated_at.isoformat()
    
    # Pretty print the XML
    xml_str = ET.tostring(root, encoding='unicode', method='xml')
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    # Remove extra blank lines
    lines = [line for line in pretty_xml.split('\n') if line.strip()]
    pretty_xml = '\n'.join(lines)
    
    return Response(
        content=pretty_xml,
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=assets.xml"}
    )


def export_to_xml_hierarchical(assets: List[Asset], metadata: dict = None) -> Response:
    """Export assets to XML format with hierarchical structure"""
    import xml.dom.minidom as minidom
    
    root = ET.Element("AssetHierarchy")
    root.set("version", settings.APP_VERSION)
    
    # Add metadata element
    if metadata:
        meta_elem = ET.SubElement(root, "ExportMetadata")
        ET.SubElement(meta_elem, "ExportDate").text = metadata['export_date']
        ET.SubElement(meta_elem, "ApplicationVersion").text = metadata['app_version']
        ET.SubElement(meta_elem, "Hostname").text = metadata['hostname']
        ET.SubElement(meta_elem, "HostIP").text = metadata['host_ip']
        ET.SubElement(meta_elem, "ClientIP").text = metadata['client_ip']
        ET.SubElement(meta_elem, "Platform").text = metadata['platform']
        ET.SubElement(meta_elem, "PythonVersion").text = metadata['python_version']
        ET.SubElement(meta_elem, "TotalAssets").text = str(len(assets))
    
    # Create a mapping of assets by ID for quick lookup
    asset_map = {asset.id: asset for asset in assets}
    
    def add_asset_to_xml(parent_elem, asset, processed=None):
        """Recursively add asset and its children to XML"""
        if processed is None:
            processed = set()
        
        if asset.id in processed:
            return
        processed.add(asset.id)
        
        asset_elem = ET.SubElement(parent_elem, "Asset")
        asset_elem.set("id", str(asset.id))
        asset_elem.set("urn", asset.urn)
        asset_elem.set("type", asset.asset_type.name)
        asset_elem.set("status", asset.status)
        
        # Basic info
        info_elem = ET.SubElement(asset_elem, "Info")
        ET.SubElement(info_elem, "Name").text = asset.name
        ET.SubElement(info_elem, "Description").text = asset.description or ""
        ET.SubElement(info_elem, "Version").text = asset.version or ""
        
        # Metadata
        if asset.properties or asset.tags or asset.external_id:
            meta_elem = ET.SubElement(asset_elem, "Metadata")
            if asset.external_id:
                ET.SubElement(meta_elem, "ExternalID").text = asset.external_id
            if asset.external_system:
                ET.SubElement(meta_elem, "ExternalSystem").text = asset.external_system
            
            # Properties
            if asset.properties:
                props_elem = ET.SubElement(meta_elem, "Properties")
                for key, value in asset.properties.items():
                    prop_elem = ET.SubElement(props_elem, "Property")
                    prop_elem.set("name", key)
                    prop_elem.text = str(value)
            
            # Tags
            if asset.tags:
                tags_elem = ET.SubElement(meta_elem, "Tags")
                for tag in asset.tags:
                    ET.SubElement(tags_elem, "Tag").text = tag
        
        # Timestamps
        timestamps_elem = ET.SubElement(asset_elem, "Timestamps")
        ET.SubElement(timestamps_elem, "Created").text = asset.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ET.SubElement(timestamps_elem, "Updated").text = asset.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        
        # Add children recursively
        children = [a for a in assets if a.parent_id == asset.id]
        if children:
            children_elem = ET.SubElement(asset_elem, "Children")
            children_elem.set("count", str(len(children)))
            for child in children:
                add_asset_to_xml(children_elem, child, processed)
    
    # Add root-level assets
    root_assets = [asset for asset in assets if asset.parent_id is None]
    root_elem = ET.SubElement(root, "RootAssets")
    root_elem.set("count", str(len(root_assets)))
    
    for asset in root_assets:
        add_asset_to_xml(root_elem, asset)
    
    # Pretty print the XML
    xml_str = ET.tostring(root, encoding='unicode', method='xml')
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ")
    
    # Remove the XML declaration added by toprettyxml and extra blank lines
    lines = pretty_xml.split('\n')
    if lines[0].startswith('<?xml'):
        lines = lines[1:]  # Remove XML declaration
    lines = [line for line in lines if line.strip()]  # Remove empty lines
    
    # Add our own XML declaration
    final_xml = '<?xml version="1.0" encoding="UTF-8"?>\n' + '\n'.join(lines)
    
    return Response(
        content=final_xml,
        media_type="application/xml",
        headers={"Content-Disposition": "attachment; filename=assets_hierarchy.xml"}
    )


def export_to_excel(assets: List[Asset]) -> Response:
    """Export assets to Excel format"""
    data = []
    for asset in assets:
        data.append({
            'ID': str(asset.id),
            'URN': asset.urn,
            'Name': asset.name,
            'Description': asset.description or '',
            'Asset Type': asset.asset_type.name,
            'Parent': asset.parent.name if asset.parent else '',
            'Status': asset.status,
            'Lifecycle Stage': asset.lifecycle_stage or '',
            'External ID': asset.external_id or '',
            'External System': asset.external_system or '',
            'Version': asset.version or '',
            'Properties': json.dumps(asset.properties),
            'Tags': ','.join(asset.tags) if asset.tags else '',
            'Created At': asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else '',
            'Updated At': asset.updated_at.strftime('%Y-%m-%d %H:%M:%S') if asset.updated_at else ''
        })
    
    df = pd.DataFrame(data)
    
    # Write to Excel with date formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl', datetime_format='YYYY-MM-DD HH:MM:SS') as writer:
        df.to_excel(writer, sheet_name='Assets', index=False)
        
        # Get the workbook and worksheet to apply additional formatting
        workbook = writer.book
        worksheet = writer.sheets['Assets']
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets.xlsx"}
    )


def export_to_excel_hierarchical(assets: List[Asset], metadata: dict = None) -> Response:
    """Export assets to Excel format with hierarchical formatting"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    data = []
    
    # Add metadata rows at the top (use -1 for Level to indicate metadata rows)
    if metadata:
        data.extend([
            {'Level': -1, 'Name': 'EXPORT METADATA', 'Asset Type': '', 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Export Date:', 'Asset Type': metadata['export_date'], 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Application Version:', 'Asset Type': metadata['app_version'], 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Host:', 'Asset Type': f"{metadata['hostname']} ({metadata['host_ip']})", 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Client IP:', 'Asset Type': metadata['client_ip'], 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Platform:', 'Asset Type': metadata['platform'], 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': 'Total Assets:', 'Asset Type': str(len(assets)), 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},
            {'Level': -1, 'Name': '', 'Asset Type': '', 'URN': '', 'Status': '', 'Description': '', 'Parent': '', 'Version': '', 'External ID': '', 'Properties': '', 'Tags': '', 'Created At': '', 'Updated At': ''},  # Empty row
        ])
    
    for asset in assets:
        # Add indentation based on hierarchy level
        level = getattr(asset, 'hierarchy_level', 0)
        indent = "  " * level  # Two spaces per level
        
        data.append({
            'Level': level,
            'Name': indent + asset.name,
            'Asset Type': asset.asset_type.name,
            'URN': asset.urn,
            'Status': asset.status,
            'Description': asset.description or '',
            'Parent': asset.parent.name if asset.parent else '',
            'Version': asset.version or '',
            'External ID': asset.external_id or '',
            'Properties': json.dumps(asset.properties) if asset.properties else '',
            'Tags': ','.join(asset.tags) if asset.tags else '',
            'Created At': asset.created_at.strftime('%Y-%m-%d %H:%M:%S') if asset.created_at else '',
            'Updated At': asset.updated_at.strftime('%Y-%m-%d %H:%M:%S') if asset.updated_at else ''
        })
    
    df = pd.DataFrame(data)
    
    # Write to Excel with formatting
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Asset Hierarchy', index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Asset Hierarchy']
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Level-based fills for hierarchy
        level_fills = [
            PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),  # Level 0
            PatternFill(start_color="F0F7FF", end_color="F0F7FF", fill_type="solid"),  # Level 1
            PatternFill(start_color="F8FBFF", end_color="F8FBFF", fill_type="solid"),  # Level 2
            None,  # Level 3+ no fill
        ]
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Apply row formatting based on hierarchy level
        metadata_fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
        metadata_font = Font(italic=True, color="8B4513")
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            level = row[0].value  # Level is in first column
            
            if level == -1:  # Metadata rows
                for cell in row:
                    cell.fill = metadata_fill
                    cell.font = metadata_font
                # Special formatting for the EXPORT METADATA header
                if "EXPORT METADATA" in str(row[1].value):
                    for cell in row:
                        cell.font = Font(bold=True, italic=True, color="8B4513")
            elif level is not None and isinstance(level, (int, float)) and level >= 0:
                level = int(level)
                if level < len(level_fills) and level_fills[level]:
                    for cell in row:
                        cell.fill = level_fills[level]
                
                # Bold font for top-level items
                if level == 0:
                    for cell in row:
                        cell.font = Font(bold=True)
        
        # Hide the Level column (column A)
        worksheet.column_dimensions['A'].hidden = True
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Set width with min and max bounds
            adjusted_width = max(10, min(max_length + 2, 50))
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze the header row
        worksheet.freeze_panes = 'A2'
        
        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows():
            for cell in row:
                cell.border = thin_border
    
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=assets_hierarchy.xlsx"}
    )